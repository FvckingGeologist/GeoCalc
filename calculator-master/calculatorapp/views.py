from itertools import count

from django.shortcuts import render
from openpyxl import load_workbook
from .models import InData2
import numpy as np
import matplotlib.pyplot as plt
import numpy_financial as npf


def home(request):
    return render(request, 'home.html')
def index(request):
    return render(request, 'index.html')

def geocalc(request):
    NullVar()
    global excel_file
    if request.method == 'POST':
#        wb = load_workbook(filename="C:\DESKTOP\INP011.xlsx", data_only=True)
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file, data_only=True)
        ws = wb["DATA"]
        for y in range(2, 700): #READ INPUT DATA A
            A[ws.cell(row=y, column=2).value] = (ws.cell(row=y, column=5).value)
        for x in range(0, 10):# MONTE-CARLO READ DATA
            TEMP = {"ID": A[321 + x], "NAME": A[331 + x], "DIM": A[341 + x], "TYPE": A[351 + x], "AVG": A[361 + x],
                    "DEV": A[371 + x], "FROM": A[381 + x], "TO": A[391 + x], "DISP": A[401 + x], "STORE": 0}
            G.append(TEMP)
        Generator()
        MonteCarlo()
        OUTPUT()
        RecDATA()
        return render(request, 'report.html', context=n)

    return render(request, 'geocalc.html')

def report(request):
    return render(request, 'report.html')

def method(request):
    return render(request, 'method.html')

def contacts(request):
    return render(request, 'contacts.html')

def help(request):
    return render(request, 'help.html')

def map(request):
    return render(request, 'map.html')

def report1(request):
    wb = load_workbook(filename="C:\DESKTOP\SAMPLE.xlsx", data_only=True)
    ws = wb.active
    n = {}
    for y in range(1, 401):
        b = ws.cell(row=y, column=2).value
        if ws.cell(row=y, column=3).value is None:
            pass
        else:
            b = [ws.cell(row=y, column=2).value]
        for x in range(3, 30):
            if ws.cell(row=y, column=x).value is None:
                pass
            else:
                b.append(ws.cell(row=y, column=x).value)
        n[ws.cell(row=y, column=1).value] = b
    return render(request, 'report.html', context=n)

def NullVar():
    global n,A,G,C,C0,B,D,D1,count1,count2,bins1
    n = {} #data for output
    A = {}  #INPUT DATA
    G = []  # INPUT DATA TO GENERATE DISTRIBUTION
    C=[] #INTERMEDIATE DATA BY YEARS
    C0=[] #INTERMEDIATE SUM|MED...
    B=[0]*100 #INTERMEDIATE CONST
    D=[] #DISTRIBUTIONS FOR MONTE_CARLO
    #D1=[[0]*10]*1000 #MONTE_CARLO_RESULT
    count1 = []
    count2 = []
    bins1=[]

def MainCalc():  # MAIN PART MAIN MAIN MAIN
    B[1] = A[3] * A[4] / 1000  # AL01 AREA CALC
    tem1 = int(A[26] + 1)  # AL02 GRR PERIOD
    B[2] = B[1] * A[2] * A[8]  # AL07 GEOL_RES_TOTAL
    B[3] = B[2] * (1 - A[28] / 100) * (1 + A[29] / 100)  # AL08 EXP_RESOURSE_TOTAL

    # AL06 OPERATION_PERIOD INCLUDE LAST YEAR (RESOURSE) ADD!!!!!!!!!!!!!!!!!!!!!!!!CONTROL
    B[15] = int(B[3] / A[31]) + 1
    B[40] = B[15]
    if (B[15] + A[26] + A[36] + 1) > A[144]:  # !!!!!!!!!!!!!!!!
        B[15] = A[144] - A[26] - A[36]  # !!!!!!!!!!!!!!!
    tem5 = A[26] + A[36] + 1 + B[15]
    B[16] = (B[15] - 1) * A[31] / B[3]  # PART OF MINED RESERVES

    # EMTY MATRIX GENERATOR
    global C
    C = [[0 for col in range(tem5)] for row in range(150)]
    global C0
    C0 = [[0 for col in range(4)] for row in range(150)]

    for x in range(1, tem1):  # NUM 1 MARKER for GRR
        C[6][x] = 1
    # AL03 GRR_VOLUME COST
    tem1 = int(A[26] + 1)
    tem = A[26]
    if A[26] == 0:
        tem = 999
    for x in range(1, tem1):
        C[1][x] = A[22]*B[1] / tem  # VOLUME
        C[2][x] = A[23]*B[1] / tem
        C[3][x] = A[24]*B[1] / tem
        C[4][x] = A[25]*B[1] / tem
        C[68][x] = (A[22] * C[1][x] + A[23] * C[2][x] + A[24] * C[3][x] + A[25] * C[4][x]) * A[148]  # GRR COST
    # AL04 GPR_PERIOD
    tem2 = int(A[26] + 1)
    tem3 = int(A[26] + A[36] + 1)
    for x in range(tem2, tem3):
        C[7][x] = 1
    # AL05 GPR_VOLUME
    tem = A[36]
    if A[36] == 0:
        tem = 999
    if A[8] == 0:
        A[8] = 0.001
    for x in range(tem2, tem3):
        C[5][x] = A[31] * A[27] / tem  # VOLUME
        C[69][x] = A[74] * C[5][x] * A[148]  # COST

    # AL09 EXP_RES_PER_YEAR
    tem4 = int(A[26] + A[36] + 1)
    for x in range(tem4, tem5):
        C[22][x] = A[31]  # EXP_RESOURSE
    if (B[3] - A[31] * (B[15] - 1)) < A[31]:  # LAST_YEAR
        C[22][tem5 - 1] = B[3] - A[31] * (B[15] - 1)
    for x in range(tem4, tem5):
        C[10][x] = C[22][x] / ((1 - A[28] / 100) * (1 + A[29] / 100))  # GEOL_RESOURSE
        C[114][x] = C[22][x] * A[39]  # OVERLAY=CAPPING
        C[133][x] = C[22][x] * (1 - A[39] / 100)  # LOSS
        C[8][x] = 1  # MARK
    for y in range(0, 9):  # EXP_CONTENT
        B[5 + y] = A[21]*A[11 + y] / (1 + A[29] / 100)

        # AL17
    B[17] = tem4
    B[18] = A[31]

    # AL20 PRODUCTION PER YEAR
    for x in range(tem4, tem5):
        for y in range(0, 9):  # for each component
            C[12 + y][x] = A[21]*A[11 + y] * C[10][x] / 100  # component 1 in geol
            C[24 + y][x] = A[12 + y] * C[22][x] / 100 / (1 + A[29] / 100)  # component in exp
            C[35 + y][x] = A[41 + y] * C[22][x]  # STAGE 1
            C[46 + y][x] = A[151 + y] * C[22][x]/A[51]
            C[57 + y][x] = A[112 + y] * C[46 + y][x] * A[147] / 1000  # PRODUCTS_PROCEEDS
            C[34][x] = C[34][x] + C[35 + y][x]  # SUMM1
            # C[45][x]=C[45][x]+C[46+y][x] NOT RECOMEND TO USE
            C[56][x] = C[56][x] + C[57 + y][x]  # PRICE
            C[116][x] = C[22][x] - C[34][x]  # PROD1 TAIL

    for x in range(1, tem5):  # AL27 CAP_INVESTMENT_PRIMARY
        delta = 0
        delta1 = 0
        #!!!!!!!!!!!!!!!! SOME BETTER SOLUTION ?
        if (C[22][x] - C[22][x - 1]) > 0:  # ORE INCREASING
            delta = (C[22][x] - C[22][x - 1])
        if (C[114][x] - C[114][x - 1]) > 0:  # CAPPING INCREASING
            delta1 = (C[114][x] - C[114][x - 1])
        C[70][x-1] = (A[80] * (delta) + A[80] * (delta1) * A[9]) * A[148]  # MINE | PIT
        C[71][x-1] = A[85] * (delta) * A[148]  # PROCEEDING1
        C[72][x-1] = A[142] * (delta) * A[148]  # TRANSPORT
        C[74][x-1] = A[91] * (delta) * A[148]  # PROCEEDING2
        C[73][x-1] = (C[70][x] + C[71][x] + C[72][x] + C[74][x]) * A[142]  # ADDITIONAL

    for x in range(1, tem5):  # AL27 CAP_INVESTMENT_PRIMARY

        C[67][x] = (C[70][x] + C[71][x] + C[72][x] + C[73][x] + C[74][x]) + C[68][x] + C[69][x]  # SUMM +GPR +GRR
        C[140][x] = C[140][x - 1] + C[67][x]  # SUMM PRIMARY COST INCLUDING GRR GPR
        C[141][x] = C[141][x - 1] + C[67][x] - C[68][x] - C[69][x]  # SAME WITHOUT GRR GPR
        C[77][x] = (C[140][x] - C[140][x - 1]) * A[98]  # CIRCULATING CAPital

    SUM = sum(C[67])
    for x in range(tem4, tem5):  # AL27 CAP_COST_ADVANSED

        # temS=temS+tem
        # tem=(C[70][x]+C[71][x]+C[72][x]+C[73][x]+C[74][x])
        C[76][x] = SUM * A[140]  # RE-INVESTMENT
        C[143][x] = C[143][x - 1] + C[76][x]  # re-investment sum

        C[145][x] = C[145][x - 1] + C[77][x]
        C[91][x] = SUM * A[99]  # DEPRICATION

#        if (C[142][x - 1] + C[91][x]) > (C[140][x] + C[143][x]):  # deprication correction
        if (C[142][x - 1] + C[91][x]) > (C[140][x]):  # OR that one

            C[91][x] = 0
        C[142][x] = C[142][x - 1] + C[91][x]  # deprication sum

    for x in range(1, tem5):  # SUMM INTERPRISE COST
        TEMP=0
        TEMP=TEMP+C[77][x]
        C[144][x] = C[140][x - 1] - C[142][x] + C[143][x] -TEMP  # sum interprise cost


        C[120][x] = C[67][x] + C[76][x]+C[77][x]  # TOTAL INVEST

    C[75][tem5 - 1] = C[144][tem5 - 1] * A[97]  # ELIMINATION
    C[120][tem5 - 1] = C[120][tem5 - 1] - C[75][tem5 - 1] - TEMP  # CORRECTION TO ELIMINATION


    # AL32 CURRENT_COST
    for x in range(tem4, tem5):
        tem = 0
        C[79][x] = A[82] * C[114][x] * A[149]  # CAPPING
        C[80][x] = A[81] * C[22][x] * A[149]  # ORE
        C[81][x] = A[38] * A[83] * C[22][x] * A[149]  # TRANSPORT
        C[82][x] = A[37] * A[83] * A[9] * C[116][x] * A[149]  # TRANSPORT
        C[83][x] = A[86] * C[22][x] * A[149]  # PROCEEDING STAGE1
        # C[84][x]=A[87]*A[51]*C[34][x]*A[149] #TRANSPORT &?????? A51
        C[85][x] = A[92] * C[22][x] * A[149]  # PROCEEDING 2
        tem = C[79][x] + C[80][x] + C[81][x] + C[82][x] + C[83][x] + C[84][x] + C[85][x]
        C[86][x] = A[94] * tem  # ADDITIONAL
        C[87][x] = A[95] * tem  # OPERATION
        C[88][x] = A[96] * tem  # ADDITIONAL2
        C[89][x] = A[97] * tem  # NATURE MANAGEMENT
        for y in range(0, 9):  # RENT
            C[90][x] = A[127 + y] * C[57 + y][x] + C[90][x]

        for y in range(0, 13):  # SUMM
            C[78][x] = C[79 + y][x] + C[78][x]
        for y in range(0, 14):  # COST PER UNIT
            if C[22][x] == 0:
                C[22][x] = 1000000000
            C[92 + y][x] = C[78 + y][x] / C[22][x]

    # AL36 MAIN ECONOMIC BLOCK
    for x in range(tem4, tem5):  # OPERATIONAL PERIOD
        C[106][x] = C[56][x] - C[78][x]  # GROSS PROFIT
        C[107][x] = C[144][x] * A[125]  # CAPITAL TAX
        C[108][x] = C[106][x] - C[107][x]  # PROFIT
        if (C[108][x] * A[123]) > 0:
            C[109][x] = C[108][x] * A[123]
            B[33] = x
        if B[33] == (tem5 - 1):
            B[33] = 999
        C[110][x] = C[108][x] - C[109][x]  # NET GAIN

    for x in range(1, tem5):
        C[118][x] = A[143] ** (x - 1)  # DISCOUNT RATE
        C[111][x] = C[90][x] + C[107][x] + C[109][x]  # STATE BUDGET PROFIT
        C[121][x] = C[110][x] + C[91][x]  # OPER BALLANCE
        C[124][x] = C[120][x]  # INVEST BALLANCE
        C[125][x] = C[121][x] - C[124][x]  # Cash Flow
        C[126][x] = C[126][x - 1] + C[125][x]  # Cash Flow SUMM
        C[127][x] = C[125][x] / C[118][x] + 0.0000001  # NPV
        if C[127][x] > 1:
            C[9][x] = 1
        C[128][x] = C[128][x - 1] + C[127][x]  # NPV SUMM
        C[129][x] = C[124][x] / C[118][x]  # Discount INvest
        C[130][x] = C[121][x] / C[118][x]  # Discount operat

    # AL43 # PERIOD OF RETURN
    B[30] = 999
    for x in range(tem4, tem5):
        if (C[127][x] > 1) and (C[127][x - 1] < 1):
            B[30] = x

    # AL50 METHOD FOR 0-SUMM 1-MED 2-MAX 3-MIN
    for x in range(0, 150):
        for y in range(0, tem5):
            C0[x][1] = C[x][y] + C0[x][1]
            if C0[x][2] < C[x][y]:
                C0[x][2] = C[x][y]
            else:
                C0[x][3] = C[x][y]
        C0[x][0] = C0[x][1] / (tem5 - tem4)

    # AL44 B_DATA_INPUT
    for x in range(0, 8):  # CAPITAL PER ORE PRODUCTION
        B[20 + x] = C0[67 + x][1] / B[18]

    B[28] = C0[127][0]  # NPV MAX
    B[29] = C0[130][0] / C0[129][0]  # PI

    CFLOW = []
    # CFLOW.clear
    for x in range(0, tem5):
        CFLOW.append(C[125][x])
    B[31] = npf.irr(CFLOW)
    B[0] = tem5
    B[39]=C0[111][1]+B[28]

def Generator():
    for x in range(0,10):
        if G[x]["TYPE"]=='Нормальное':
            mu, sigma = 1., G[x]['DEV']/G[x]['AVG']
            TEMP=[]
            TEMP=np.random.normal(mu, sigma, 1000)
            for y in range(0,1000):
                TEMP[y]=G[x]['AVG']*TEMP[y]
            D.append(TEMP)
        else:
            D.append([0]*1000)

def MonteCarlo():
    TEMP = []
    TEMP1 = 0
    STEP=0

    for y in range(0, 10):  # store values to temp
        if G[y]['ID'] > 0:
            G[y]["STORE"] = A[G[y]['ID']]

    for x in range(1, 999):
        for y in range(0, 10):  # Monte-carlo function
            if G[y]['ID'] > 0:
                A[G[y]['ID']] = D[y][x]
                # D1[x][y] = A[G[y]['ID']] ??????????????? WTF7
        MainCalc()
        TEMP.append(B[28])

    for y in range(0, 10):  # return values to data
        if G[y]['ID'] > 0:
            A[G[y]['ID']] = G[y]["STORE"]
    MainCalc()
    TEMP.sort()
    # D.append(TEMP) #service check

    for x in range(0, 998):  # probability calculation
        if TEMP[x] > 0:
            B[32] = x / 999 * 100
            x=1000

    STEP=(TEMP[950]-TEMP[50])/24 #HYSTOGRAMM SYBUNIT
    y=0
    xo=0
    for x in range(50, 950):
            if TEMP[x] > (STEP*y+TEMP[50]):
                bins1.append(round(y*STEP/1000,0))
                y=y+1
                if TEMP[x]>0:
                    count1.append(x-xo)
                    count2.append(0)
                else:
                    count2.append(x-xo)
                    count1.append(0)
                xo=x


def OUTPUT():
    n.update({
        'N1': round(B[32], 0),#Round1
        'N2': round(B[39]/1000, 0),#Table1
        'N3': round(B[28]/1000, 0),
        'N4': round(B[29], 2),
        'N5': round(B[31] * 100, 2),
        'N6': round(C[111][1]/1000, 0),

        'N7': 0,
    })
    for x in range(0, 7):
        n['N' + str(8 + x)] = A[501 + x]
    n.update({
        'N15': C0[114][1],
        'N16': C0[22][1],
        'N17': 0,
        'N18': 0,
        'N19': 0
    })
    for x in range(0, 7):
        n['N' + str(20 + x)] = A[508 + x]
    n.update({
        'N27': C0[10][1],
        'N28': C0[133][1],
        'N29': C0[22][1]-C0[10][1],
        'N30': 0,
        'N31': 0
    })
    for x in range(0, 7):
        n['N' + str(32 + x)] = A[515 + x]
    n.update({
        'N39': C0[34][1],
        'N40': 0,
        'N41': C0[116][1],
        'N42': 0,
        'N43': 0
    })
    for x in range(0, 8):
        n['N' + str(44 + x)] = A[522 + x]
    n.update({
        'N52': C[22],
        'N53': C[114],
        'N54': C[5],
        'N55': 0,
        'N56': 0
    })
    for x in range(0, 21):
        n['N' + str(57 + x)] = A[530 + x]
    n.update({
        'N78': round(B[2], 0),
        'N79': round(B[3], 0),
        'N80': round(B[3]+C0[114][1], 0),
        'N81': round(A[28], 1),
        'N82': round(A[29], 1),
        'N83': round(B[40], 0),
        'N84': round(B[18], 0),
        'N85': round(B[17], 0),
        'N86': round(A[39], 0),
        'N87': round(C0[34][1], 0)
    })
    for x in range(0, 14):
        n['N' + str(88 + x)] = A[551 + x]

    TEMP1=[]
    TEMP2 = []
    TEMP3 = []
    TEMP4 = []
    TEMP5 = []
    TEMP6 = []
    TEMP7 = []
    TEMP8 = []
    TEMP9 = []
    TEMP10 = []

    for x in range(0, B[15]+B[17]):

        TEMP1.append(-C[68][x]/1000)
        TEMP2.append(-C[69][x]/1000)
        TEMP3.append((-C[67][x]+C[68][x]-C[69][x])/1000)
        TEMP4.append(-C[76][x]/1000)
        TEMP5.append((-C[78][x] + C[91][x])/1000)
        TEMP6.append(-C[111][x]/1000)
        TEMP7.append(C[91][x] / 1000)
        TEMP8.append(C[109][x] / 1000)
        TEMP9.append(C[128][x] / 1000)
    n.update({
        'N101': TEMP1,
        'N102': TEMP2,
        'N103': TEMP3,
        'N104': TEMP4,
        'N105': TEMP5,
        'N106': TEMP6,
        'N107': TEMP7,
        'N108': TEMP8,
        'N109': TEMP9,
        'N110': 0
    })
    for x in range(0, 21):
        n['N' + str(111 + x)] = A[564 + x]
    n.update({
        'N132': round(C0[56][1]/1000,0),
        'N133': round(C0[67][1]/1000,0),
        'N134': round(C0[77][1]/1000,0),
        'N135': round(C0[78][1]/1000,0),
        'N136': round(C0[106][1]/1000,0),
        'N137': round(C0[107][1]/1000,0),
        'N138': round(C0[108][1]/1000,0),
        'N139': round(C0[109][1]/1000,0),
        'N140': round(C0[110][1]/1000,0),
        'N141': round(B[28]/1000,0)
    })
    for x in range(0, 15):
        n['N' + str(142 + x)] = A[585 + x]
    for x in range(0, 13):
        n['N' + str(157 + x)] = round(C0[79 + x][1]/1000,0)
    for x in range(0, 12):
        n['N' + str(170 + x)] = A[600 + x]
    for x in range(0, 10):
        n['N' + str(182 + x)] = round(C0[68 + x][1]/1000,0)
    for x in range(0, 27):
        n['N' + str(192 + x)] = A[612 + x]
    for x in range(0, 13):
        n['N' + str(219 + x)] = round(C0[79 + x][1]/1000,0)
    for x in range(0, 21):
        n['N' + str(232 + x)] = A[639 + x]
    for x in range(0, 10):
        n['N' + str(253 + x)] = round(C0[68 + x][1]/1000,0)
    for x in range(0, 81):
        n['N' + str(263 + x)] = A[331 + x]
    n.update({
        'N343': count1,
        'N344': count2,
        'N345': bins1,
        'N346': A[10],
        #'N347': A[261]+A[262]+A[263]+'...',
        'N347': D,
        'N349': G

    })
def RecDATA():
    if np.isnan(B[31]):
        B[31]=-1
    InD = InData2(DESCRIPTION=A[10], NPV=B[28], IRR=B[31], PI=B[29], SE=B[39], BP=C0[111][1], RESERVES=B[2],
                  PROD_RATE=A[31], ORE_TYPE=A[261] + A[262] + A[263])
#    InD=InData2( DESCRIPTION=A[10], NPV=B[28], IRR=B[31], PI=B[29], SE=B[39], BP=C0[111][1], RESERVES=B[2], PROD_RATE=A[31], ORE_TYPE=A[261]+A[262]+A[263], STORE_FILE=excel_file)
    InD.save()

def Save_XLS():
    wb = load_workbook(excel_file, data_only=True)
    ws1 = wb["SERV1"]
    ws2 = wb["SERV2"]

    tem5 = A[26] + A[36] + 1 + B[15]

    for x in range(0, 145):
        for y in range(0, 4):
            ws1.cell(row=x + 2, column=y + 4).value = C0[x][y]
        for y in range(0, tem5):
            ws1.cell(row=x + 2, column=y + 8).value = C[x][y]
    for x in range(0, 49):
        ws2.cell(row=x + 2, column=4).value = B[x]

#    for x in range(1, 330): #Dictionary output (report form)
#        ws2.cell(row=x + 1, column=9).value = n['N'+str(x)] ADD LIST RULE

    wb.save(excel_file)